from openpyxl import load_workbook

wb = load_workbook("scores.xlsx", data_only=True)
ws = wb.active

# for cols in ws.values:
#     for cell in cols:
#         print(cell)
for i in range(2, 12):
    if ws["H{}".format(i)].value >= 90:
        ws["I{}".format(i)] = "A"
    elif 90 > ws["H{}".format(i)].value >= 80:
        ws["I{}".format(i)] = "B"
    elif 80 > ws["H{}".format(i)].value >= 70:
        ws["I{}".format(i)] = "C"
    else:
        ws["I{}".format(i)] = "D"

for j in range(2, 12):
    if ws["B{}".format(j)].value < 5:
        ws["I{}".format(j)] = "F"

wb.save("scores.xlsx")
