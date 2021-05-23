from openpyxl import Workbook, load_workbook

wb = Workbook()
wb2 = load_workbook("scores.xlsx", data_only=True)
ws = wb.active

header = ["Student ID", "Attendance", "Quiz 1", "Quiz 2", "Mid Term", "Final", "Project", "Total", "score"]
ws.append(header)

scores = [
    [1,10,8,5,14,26,12], 
    [2,7,3,7,15,24,18], 
    [3,9,5,8,8,12,4], 
    [4,7,8,7,17,21,18], 
    [5,7,8,7,16,25,15], 
    [6,3,5,8,8,17,0], 
    [7,4,9,10,16,27,18], 
    [8,6,6,6,15,19,17], 
    [9,10,10,9,19,30,19], 
    [10,9,8,8,20,25,20]
    
    ]
for x in scores:
    ws.append(x)

for rows in ws.iter_rows(min_row=2):
    rows[3].value = 10
    # print(rows[3].value)

# ws["H2"] = "=SUM(B2:G2)"
# ws["H3"] = "=SUM(B3:G3)"
# ws["H4"] = "=SUM(B4:G4)"
# ws["H5"] = "=SUM(B5:G5)"
# ws["H6"] = "=SUM(B6:G6)"
# ws["H7"] = "=SUM(B7:G7)"
# ws["H8"] = "=SUM(B8:G8)"
# ws["H9"] = "=SUM(B9:G9)"
# ws["H10"] = "=SUM(B10:G10)"
# ws["H11"] = "=SUM(B11:G11)"

for i in range(2, 12):
    ws["H{}".format(i)] = ("=SUM(B{}:G{})".format(i, i))

for rows in ws.iter_rows(min_row=2):
    # print(rows[7].value)
    pass


wb.save("scores.xlsx")