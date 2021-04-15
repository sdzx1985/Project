from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if row[1].value > 90:
        print(row[0].value)

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "English":
            cell.value = "Computer"

wb.save("sample_modi.xlsx")