from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

ws.append(["Number", "English", "Math"])
for i in range(1, 11): 
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"]
# print(col_B)
for cell in col_B:
    print(cell.value)

col_range = ws["B:C"]


wb.save("sample.xlsx")