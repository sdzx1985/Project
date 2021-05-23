from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

ws.append(["Number", "English", "Math"])
for i in range(1, 11): 
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # call english column only
# print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"] # call english and math column
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1] # call 1st row 
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # call from 2nd to 6th row / diff with slice(upto 6th)
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] # call from 2nd to the last row
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end = " ")
#         # print(cell.coordinate, end=" ") # call the cell's info (name, location)
#         xy = coordinate_from_string(cell.coordinate) # make it as tuple
#         # print(xy, end=" ")
#         print(xy[0], end="") # A
#         print(xy[1], end=" ") # 1
#     print()

# all rows
# print (tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[0].value)

# all colums
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows(): # all rows
#     print(row[1].value)

# for column in ws.iter_cols(): # all columns
#     print (column[0].value)

# from 1st row to 5th row, 2dn columns to 3rd columns (call data horizontal)
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3): 
    # print(row[0].value, row[1].value)
    # print(row)

# call data vertical
for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

wb.save("sample.xlsx")