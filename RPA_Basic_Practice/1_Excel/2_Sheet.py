from openpyxl import Workbook
wb = Workbook()
#wb.active
ws = wb.create_sheet() # create a new sheet / basic name
ws.title = "MySheet" # change sheet name
ws.sheet_properties.tabColor = "ff66ff" #RGB

ws1  = wb.create_sheet("YourSheet") # create a new sheet / given name
ws2 = wb.create_sheet("NewSheet", 2) # create a new sheet / 2nd index


new_ws = wb["NewSheet"] # access the sheet with dict method

print(wb.sheetnames) # check the all sheet name

# copy sheet
new_ws["A1"] = "Test" # Insert "Test" at the A1 cell
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
