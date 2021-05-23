from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# number, english, math to number, korean, english, math
# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "Korean"

ws.move_range("C1:C11", rows=5, cols=-1)

wb.save("Sample_Korean.xlsx")